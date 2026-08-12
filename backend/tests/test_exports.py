"""Tests for Excel and HTML report export endpoints."""
import pytest
from httpx import AsyncClient, ASGITransport
from sqlalchemy import select

from app.main import app
from app.models import Audit, Issue, Page, Recommendation, User
from app.database import async_session

transport = ASGITransport(app=app)


async def _auth_headers(client, email, username=None):
    username = username or email.split("@")[0]
    await client.post("/api/auth/register", json={
        "email": email, "username": username, "password": "password123",
    })
    resp = await client.post("/api/auth/login", json={
        "email": email, "password": "password123",
    })
    return {"Authorization": f"Bearer {resp.json()['access_token']}"}


async def _user_id(email: str) -> str:
    async with async_session() as db:
        user = (await db.execute(select(User).where(User.email == email))).scalar_one()
        return user.id


async def _create_audit_with_data(user_id: str) -> str:
    async with async_session() as db:
        audit = Audit(website_url="https://export-test.com", status="COMPLETED", progress=100, user_id=user_id)
        db.add(audit)
        await db.flush()
        db.add_all([
            Issue(audit_id=audit.id, page_url="https://export-test.com/", category="TECHNICAL",
                  severity="CRITICAL", signal_name="Missing Title",
                  description="No title", impact="High", fix="Add a title"),
            Page(audit_id=audit.id, url="https://export-test.com/", status_code=200,
                 title="Home", word_count=500, page_type="HOMEPAGE"),
            Recommendation(audit_id=audit.id, page_url="https://export-test.com/", category="TECHNICAL",
                           priority="HIGH", issue="Missing Title", current_problem="No title",
                           why_it_matters="Ranking", exact_fix="Add title", difficulty="EASY",
                           expected_impact="High"),
        ])
        await db.commit()
        return audit.id


@pytest.mark.asyncio
async def test_excel_export_full():
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        headers = await _auth_headers(client, "export_user@test.com")
        audit_id = await _create_audit_with_data(await _user_id("export_user@test.com"))

        resp = await client.get(f"/api/audit/{audit_id}/export/excel", headers=headers)
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        assert f"seo-report-{audit_id[:8]}.xlsx" in resp.headers.get("content-disposition", "")

        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(resp.content))
        assert set(wb.sheetnames) == {"Overview", "Issues", "Pages", "Recommendations"}
        ws = wb["Issues"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("Page URL", "Category", "Severity", "Issue", "Description", "Impact", "Fix")
        assert any("Missing Title" in str(r) for r in rows)


@pytest.mark.asyncio
async def test_excel_export_single_type():
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        headers = await _auth_headers(client, "export_user2@test.com")
        audit_id = await _create_audit_with_data(await _user_id("export_user2@test.com"))

        resp = await client.get(f"/api/audit/{audit_id}/export/excel", params={"type": "issues"}, headers=headers)
        assert resp.status_code == 200
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(resp.content))
        assert "Issues" in wb.sheetnames
        assert "Pages" not in wb.sheetnames


@pytest.mark.asyncio
async def test_excel_export_bad_type():
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        headers = await _auth_headers(client, "export_user3@test.com")
        audit_id = await _create_audit_with_data(await _user_id("export_user3@test.com"))
        resp = await client.get(f"/api/audit/{audit_id}/export/excel", params={"type": "nope"}, headers=headers)
        assert resp.status_code == 400


@pytest.mark.asyncio
async def test_html_export_contains_report():
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        headers = await _auth_headers(client, "export_user4@test.com")
        audit_id = await _create_audit_with_data(await _user_id("export_user4@test.com"))

        resp = await client.get(f"/api/audit/{audit_id}/export/html", headers=headers)
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith("text/html")
        assert "export-test.com" in resp.text
        assert "Critical Issues" in resp.text
        assert "AI SEO Intelligence Report" in resp.text
        assert "Missing Title" in resp.text


@pytest.mark.asyncio
async def test_export_audit_not_found():
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        headers = await _auth_headers(client, "export_user5@test.com")
        resp = await client.get("/api/audit/no-such/export/excel", headers=headers)
        assert resp.status_code == 404
        resp = await client.get("/api/audit/no-such/export/html", headers=headers)
        assert resp.status_code == 404


@pytest.mark.asyncio
async def test_update_google_properties():
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        headers = await _auth_headers(client, "export_user6@test.com")
        audit_id = await _create_audit_with_data(await _user_id("export_user6@test.com"))
        resp = await client.put(f"/api/audit/{audit_id}/google-properties", json={
            "gsc_property": "sc-domain:example.com",
            "ga_property": "properties/123456",
        }, headers=headers)
        assert resp.status_code == 200
        assert resp.json()["ga_property"] == "properties/123456"


@pytest.mark.asyncio
async def test_update_google_properties_requires_auth():
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        resp = await client.put("/api/audit/x/google-properties", json={"gsc_property": "", "ga_property": ""})
        assert resp.status_code == 401


@pytest.mark.asyncio
async def test_ga4_properties_requires_connected_account():
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        headers = await _auth_headers(client, "export_user7@test.com")
        resp = await client.get("/api/integrations/google/ga4-properties", headers=headers)
        assert resp.status_code == 404
